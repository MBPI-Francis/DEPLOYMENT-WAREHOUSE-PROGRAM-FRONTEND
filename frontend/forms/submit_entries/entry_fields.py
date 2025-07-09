import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import requests
from backend.settings.database import server_ip
from ttkbootstrap.tooltip import ToolTip
from ttkbootstrap.dialogs.dialogs import Messagebox
from datetime import datetime, timedelta
from .table import SubmitEntriesTable
from tkinter.filedialog import asksaveasfilename
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import numbers
from frontend.forms.shared import SharedFunctions
from frontend.constant import FormConstant
import os
import time # Import for potential small delay

# Try to import win32com.client for full workbook encryption
try:
    import win32com.client as win32
    WIN32_AVAILABLE = True
except ImportError:
    WIN32_AVAILABLE = False
    # Messagebox.show_warning(
    #     "pywin32 library not found. Full workbook encryption will not be available. "
    #     "Only sheet protection will be applied (if password is set). "
    #     "Install with: pip install pywin32",
    #     "Library Missing"
    # )

excel_password = FormConstant.EXCEL_SOH_PASSWORD.value


def entry_fields(note_form_tab):
    shared_functions = SharedFunctions()

    def export_data_to_excel():

        date_entry_value = date_entry.entry.get()

        # Convert date to YYYY-MM-DD
        try:
            date_entry_value = datetime.strptime(date_entry_value, "%m/%d/%Y").strftime("%Y-%m-%d")
        except ValueError:
            Messagebox.show_error("Invalid date format. Please use MM/DD/YYYY.", "Date Entry Error")
            return

        try:

            # Call the function to get all the data from the view table
            data = get_soh_data()

            # Generate an Excel file for the stock-on-hand data
            create_soh_whse_excel(date_entry_value, data)

        except requests.exceptions.RequestException as e:
            # Show an error message if the first POST request fails
            Messagebox.show_info(e, "Data Entry Error")


    # Function to update stock
    def update_stock():
        date_entry_value = date_entry.entry.get()

        # Convert date to YYYY-MM-DD
        try:
            date_entry_value = datetime.strptime(date_entry_value, "%m/%d/%Y").strftime("%Y-%m-%d")
        except ValueError:
            Messagebox.show_error("Invalid date format. Please use MM/DD/YYYY.", "Date Entry Error")
            return

        try:
            # Send another POST request to update the stocks
            response = requests.post(f"{server_ip}/api/update_stock_on_hand/?params_date={date_entry_value}")
            if response.status_code == 200:
                # print("Successfully Updated the Stocks.")

                try:
                    # Send a POST request to update the computed date in the API
                    response = requests.post(f"{server_ip}/api/update-date-computed")
                    if response.status_code == 200:
                        # print("Successfully Updated the Computed Date")
                        pass
                except requests.exceptions.RequestException as e:
                    Messagebox.show_error(e, "Data Entry Error")


                # Refresh the note table to display updated data
                note_table.refresh_table()

                # Show success message
                Messagebox.show_info("The new beginning balance has been successfully updated!", "Success")

        except requests.exceptions.RequestException as e:
            Messagebox.show_info(e, "Data Entry Error")


    # Function to show confirmation panel
    def show_confirmation_panel():
        # confirmation_window = ttk.Toplevel(form_frame)
        # confirmation_window.title("Confirm Action")
        # confirmation_window.geometry("490x290")
        # confirmation_window.resizable(False, False)

        confirmation_window = ttk.Toplevel(form_frame)
        confirmation_window.title("Confirm Action")

        # Get the screen width and height
        screen_width = confirmation_window.winfo_screenwidth()
        screen_height = confirmation_window.winfo_screenheight()

        # Set a dynamic size (proportional to the screen size)
        window_width = int(screen_width * 0.42)  # Adjust width as needed
        window_height = int(screen_height * 0.43)  # Adjust height as needed

        # Calculate position for centering
        x_position = (screen_width - window_width) // 2
        y_position = (screen_height - window_height) // 3  # Position slightly higher

        # Apply geometry dynamically
        confirmation_window.geometry(f"{window_width}x{window_height}+{x_position}+{y_position}")

        # Allow resizing but maintain proportions
        confirmation_window.resizable(True, True)

        # Expand and fill widgets inside the window
        confirmation_window.grid_columnconfigure(0, weight=1)
        confirmation_window.grid_rowconfigure(0, weight=1)


        # Message Label
        message_label = ttk.Label(
            confirmation_window,
            text="\n\nYou are about to set the new beginning balance",
            justify="center",
            font=("Arial", 14, "bold"),
            bootstyle=WARNING

        )
        message_label.pack(pady=5)

        # Message Label
        message_label = ttk.Label(
            confirmation_window,
            text=(
                "Please review all the data you have entered before proceeding\n\n"
                "Setting a new beginning balance is a permanent action and cannot be undone.\n"
                "Ensure that all values are correct before confirming.\n"
            ),
            justify="center",
            font=("Arial", 11),
        )
        message_label.pack(pady=0)

        # Message Label
        message_label = ttk.Label(
            confirmation_window,
            text=(
                f"Note: This action will also record the Inventory Report Date as {date_entry.entry.get()}\n"
            ),
            justify="center",
            font=("Arial", 11, "bold"),
        )
        message_label.pack(pady=0)


        # Message Label
        message_label = ttk.Label(
            confirmation_window,
            text=(
                "To proceed, type 'YES' in the confirmation box."
            ),
            justify="center",
            font=("Arial", 11, 'bold'),
        )
        message_label.pack(pady=0)


        # Entry field
        confirm_entry = ttk.Entry(confirmation_window, font=("Arial", 12),
                                  justify = "center")
        confirm_entry.pack(padx=20, pady=5)

        # Frame for buttons
        button_frame = ttk.Frame(confirmation_window)
        button_frame.pack(fill="x", padx=10, pady=10)  # Expand the frame horizontally

        # Configure button frame columns
        button_frame.columnconfigure(0, weight=1)  # Left side (Cancel)
        button_frame.columnconfigure(1, weight=1)  # Right side (Submit)

        # Cancel Button (Left)
        cancel_button = ttk.Button(
            button_frame,
            text="Cancel",
            bootstyle=DANGER,
            command=confirmation_window.destroy
        )
        cancel_button.grid(row=0, column=0, padx=5, sticky="w")  # Align to left

        # Submit Button (Right, Initially Disabled)
        submit_button = ttk.Button(
            button_frame,
            text="Submit",
            bootstyle=SUCCESS,
            state=DISABLED,
            command=lambda: [update_stock(), confirmation_window.destroy()]
        )
        submit_button.grid(row=0, column=1, padx=5, sticky="e")  # Align to right

        # Function to validate entry field
        def validate_entry(event):
            if confirm_entry.get().strip() == "YES":
                submit_button.config(state=NORMAL)
            else:
                submit_button.config(state=DISABLED)

        confirm_entry.bind("<KeyRelease>", validate_entry)



    def format_date_while_typing(event):
        """Auto-formats the date entry while typing, ensuring valid MM/DD/YYYY format."""
        text = date_entry.entry.get().replace("/", "")  # Remove slashes to get raw number input
        formatted_date = ""

        if len(text) > 8:  # Prevent overflow of more than 8 characters (MMDDYYYY)
            text = text[:8]

        # Handle the case where the length is 2 (MD)
        if len(text) == 2:
            month = text[:1]
            day = text[1:]
            # Ensure the month is valid (01-12)
            if int(month) < 1 or int(month) > 12:
                month = str(datetime.now().month)  # Default to January if invalid month
            # Ensure the day is valid (01-31)
            if int(day) < 1 or int(day) > 31:
                day = str(datetime.now().day)  # Default to 1st if invalid day
            year = str(datetime.now().year)  # Assume the current year
            formatted_date = f"0{month}/0{day}/{year}"


        # Handle the case where the length is 3 (MDD)
        elif len(text) == 3:
            month = text[:1]
            day = text[1:]
            # Ensure the month is valid (01-12)
            if int(month) < 1 or int(month) > 12:
                month = str(datetime.now().month)  # Default to January if invalid month
            # Ensure the day is valid (01-31)
            if int(day) < 1 or int(day) > 31:
                day = str(datetime.now().day)  # Default to 1st if invalid day
                if len(str(day)) == 1:
                    day = f"0{day}"

            year = str(datetime.now().year)  # Assume the current year
            formatted_date = f"0{month}/{day}/{year}"


        # Handle the case where the length is 4 (MMDD)
        elif len(text) == 4:
            month = text[:2]
            day = text[2:]
            # Ensure the month is valid (01-12)
            if int(month) < 1 or int(month) > 12:
                month = str(datetime.now().month)  # Default to January if invalid month
                if len(str(month)) == 1:
                    month = f"0{month}"

            # Ensure the day is valid (01-31)
            if int(day) < 1 or int(day) > 31:
                day = str(datetime.now().day)  # Default to 1st if invalid day
                if len(str(day)) == 1:
                    day = f"0{day}"

            year = str(datetime.now().year)  # Assume the current year
            formatted_date = f"{month}/{day}/{year}"

            # Handle the case where the length is 5 (MDDYY)
        elif len(text) == 5:
            month = text[:1]
            day = text[1:3]
            year = text[3:]

            # Ensure the month is valid (01-12)
            if int(month) < 1 or int(month) > 12:
                month = str(datetime.now().month)  # Default to January if invalid month
            # Ensure the day is valid (01-31)
            if int(day) < 1 or int(day) > 31:
                day = str(datetime.now().day)  # Default to 1st if invalid day
                if len(str(day)) == 1:
                    day = f"0{day}"
            formatted_date = f"0{month}/{day}/20{year}"

        # Handle the case where the length is 6 (MMDDYY)
        elif len(text) == 6:
            month = text[:2]
            day = text[2:4]
            year = text[4:]
            # Ensure the month is valid (01-12)
            if int(month) < 1 or int(month) > 12:
                month = str(datetime.now().month)  # Default to January if invalid month
                if len(str(month)) == 1:
                    month = f"0{month}"

            # Ensure the day is valid (01-31)
            if int(day) < 1 or int(day) > 31:
                day = str(datetime.now().day)  # Default to 1st if invalid day
                if len(str(day)) == 1:
                    day = f"0{day}"

            formatted_date = f"{month}/{day}/20{year}"  # Assume 20XX for 2-digit year


            # Handle the case where the length is 6 (MMDDYY)
        elif len(text) == 7:
            Messagebox.show_error("Invalid date format. Please use MM/DD/YYYY.", "Date Entry Error")


        # Handle the case where the length is 8 (MMDDYYYY)
        elif len(text) == 8:
            month = text[:2]
            day = text[2:4]
            year = text[4:]
            # Ensure the month is valid (01-12)
            if int(month) < 1 or int(month) > 12:
                month = str(datetime.now().month)  # Default to January if invalid month
                if len(str(month)) == 1:
                    month = f"0{month}"

            # Ensure the day is valid (01-31)
            if int(day) < 1 or int(day) > 31:
                day = str(datetime.now().day)  # Default to 1st if invalid day
                if len(str(day)) == 1:
                    day = f"0{day}"

            formatted_date = f"{month}/{day}/{year}"

        # Update entry field with formatted value
        date_entry.entry.delete(0, "end")
        date_entry.entry.insert(0, formatted_date)

    # Create a frame for the form inputs
    form_frame = ttk.Frame(note_form_tab)
    form_frame.pack(fill=X, pady=10, padx=20)


    # Quantity Entry Field
    date_label = ttk.Label(form_frame, text="Ending Inventory Report Date", style="CustomLabel.TLabel")
    date_label.grid(row=1, column=0, padx=2, pady=(0, 0), sticky=W)

    # Calculate yesterday's date
    yesterday_date = datetime.now() - timedelta(days=1)

    # Create the DateEntry widget with yesterday's date as the default value
    date_entry = ttk.DateEntry(
        form_frame,
        bootstyle=PRIMARY,
        dateformat="%m/%d/%Y",
        startdate=yesterday_date,  # Set yesterday's date
        width=30
    )
    date_entry.grid(row=2, column=0, padx=5, pady=(0,0), sticky=W)
    date_entry.entry.config(font=shared_functions.custom_font_size)
    ToolTip(date_entry, text=f"This is the default consumption date. You can manually change it.")
    date_entry.entry.bind("<FocusOut>", format_date_while_typing)
    date_entry.entry.bind("<Return>", format_date_while_typing)


    # Add button to export data into excel
    btn_export = ttk.Button(
        form_frame,
        text="EXPORT TO EXCEL",
        command=export_data_to_excel,
        bootstyle=SUCCESS
    )
    btn_export.grid(row=2, column=2, pady=(0,0))
    ToolTip(btn_export, text="Click the button to export the data into excel.")


    # Add button to submit data
    btn_submit = ttk.Button(
        form_frame,
        text="MAKE THIS DATA AS THE NEW BEGINNING BALANCE",
        command=show_confirmation_panel,
        bootstyle=INFO
    )
    btn_submit.grid(row=2, column=3, pady=(0,0),padx=5,)
    ToolTip(btn_submit, text="Click the button to make this data as the new beginning balance")


    # Calling the table
    note_table = SubmitEntriesTable(note_form_tab)

# def create_soh_whse_excel(date_entry_value, data):
#     # Convert the string into a datetime object
#     notes_date_object = datetime.strptime(date_entry_value, "%Y-%m-%d")
#     notes_formatted_date = notes_date_object.strftime("%B %d, %Y")
#     notes_date = notes_formatted_date
#
#     wh_date_object = datetime.strptime(date_entry_value, "%Y-%m-%d")
#     wh_formatted_date = wh_date_object.strftime("%m/%d/%Y")
#     wh_date = wh_formatted_date
#
#     # Create a new workbook
#     wb = Workbook()
#
#     # Sheet 1: NOTES
#     notes_sheet = wb.active
#     notes_sheet.title = "NOTES"
#
#     # Populate the NOTES sheet
#     notes_sheet["A1"] = "Daily Ending Inventory Report from:"
#     notes_sheet["B1"] = f"{notes_date}"  # Sample date
#     notes_sheet["A2"] = "List of Batches Included in Report"
#     notes_sheet["A3"] = "MASTERBATCH"
#     notes_sheet.append(["PRODUCT CODE", "LOT#", "Product Kind"])
#
#     # Fetch data from the API
#     try:
#         api_url = f"{server_ip}/api/notes/v1/list/"
#         response = requests.get(api_url)
#         response.raise_for_status()
#         api_data = response.json()
#
#         # Insert data into the NOTES sheet
#         for record in api_data:
#             notes_sheet.append([
#                 record["product_code"],
#                 record["lot_number"],
#                 record["product_kind_id"]
#             ])
#
#     except requests.exceptions.RequestException as e:
#         # print(f"Error fetching data from API: {e}")
#         return
#
#     # Apply formatting
#     for col in ["A", "B", "C"]:
#         for cell in notes_sheet[col]:
#             cell.alignment = Alignment(horizontal="center", vertical="center")
#     notes_sheet["A4"].font = Font(bold=True)
#
#     # Function to create a WHSE sheet
#     def create_whse_sheet(sheet_name):
#         sheet = wb.create_sheet(sheet_name)
#
#         # Set header based on warehouse number
#         if sheet_name == "WHSE1":
#             wh_header = "WHSE #1 - Excess"
#         elif sheet_name == "WHSE2":
#             wh_header = "WHSE #2 - Excess"
#         elif sheet_name == "WHSE4":
#             wh_header = "WHSE #4 - Excess"
#
#         # Populate the header
#         header = [
#             "Date", "No of bags", "qty per packing",
#             f"{wh_header}", "Total", "Status"
#         ]
#         sheet.append(header)
#         sheet["A1"] = f"{wh_date}"  # Example date
#         sheet["A1"].font = Font(bold=True)
#
#         # Insert data into the respective warehouse sheet
#         for record in data:
#             if record["warehousenumber"] == int(sheet_name[-1]):  # Match warehouse number to sheet
#
#                 row = [
#                     record["rmcode"],  # rmcode
#                     "",  # No of bags (blank)
#                     "",  # qty per packing (blank)
#                     "",  # Excess column (blank)
#                     float(record["new_beginning_balance"]),  # Total
#                     "" if record["status"].lower() == "good" else record["status"],  # Store blank if status is 'good'
#                     ""  # Drop list (blank)
#                 ]
#                 sheet.append(row)
#
#                 # Apply number format (thousands separator)
#                 qty_cell = sheet.cell(row=sheet.max_row, column=5)  # Column 5 is qty_value
#                 qty_cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1  # Apply Excel formatting
#
#         # Create a dropdown list for the "drop list" column
#         dv = DataValidation(
#             type="list",
#             formula1='"held : under evaluation,held : reject,held : contaminated"',
#             allow_blank=True,
#             showDropDown=True
#         )
#         # Apply the data validation to the G column for rows 2 to 100
#         for row in range(2, 101):
#             cell = f"G{row}"  # Example: G2, G3, ..., G100
#             dv.add(sheet[cell])
#         sheet.add_data_validation(dv)
#
#     # Create sheets for WHSE1, WHSE2, and WHSE4
#     create_whse_sheet("WHSE1")
#     create_whse_sheet("WHSE2")
#     create_whse_sheet("WHSE4")
#
#
#     # Use tkinter's asksaveasfilename for file dialog
#     file_path = asksaveasfilename(
#         title="Save Excel File",
#         defaultextension=".xlsx",
#         filetypes=[("Excel files", "*.xlsx")]
#     )
#
#     # Save the workbook
#     if file_path:
#         try:
#             wb.save(file_path)
#             Messagebox.show_info(f"Excel file saved successfully at {file_path}", "The File is Successfully Saved")
#         except Exception as e:
#             Messagebox.show_error(f"Error saving file: {e}", "Success")
#
#     else:
#         # Saving the file was cancelled
#         pass
#





def create_soh_whse_excel(date_entry_value, data):
    """
    Exports raw material movement data to an Excel file with multiple sheets.
    Encrypts the entire workbook with a password if on Windows with Excel installed.

    Args:
        date_entry_value (str): The date string in "YYYY-MM-DD" format.
        data (list): A list of dictionaries containing raw material movement data.
                     Each dictionary should have 'warehousenumber', 'rmcode',
                     'new_beginning_balance', and 'status'.
        excel_password (str, optional): The password to encrypt the entire Excel workbook.
                                        If provided, the file will require this password to open.
                                        Defaults to None (no workbook encryption).
    """
    # Convert the string into datetime objects
    notes_date_object = datetime.strptime(date_entry_value, "%Y-%m-%d")
    notes_formatted_date = notes_date_object.strftime("%B %d, %Y")
    notes_date = notes_formatted_date

    wh_date_object = datetime.strptime(date_entry_value, "%Y-%m-%d")
    wh_formatted_date = wh_date_object.strftime("%m/%d/%Y")
    wh_date = wh_formatted_date

    # Create a new workbook
    wb = Workbook()

    # Sheet 1: NOTES
    notes_sheet = wb.active
    notes_sheet.title = "NOTES"

    # Populate the NOTES sheet
    notes_sheet["A1"] = "Daily Ending Inventory Report from:"
    notes_sheet["B1"] = f"{notes_date}"
    notes_sheet["A2"] = "List of Batches Included in Report"
    notes_sheet["A3"] = "MASTERBATCH"
    notes_sheet.append(["PRODUCT CODE", "LOT#", "Product Kind"])

    # Fetch data from the API
    try:
        api_url = f"{server_ip}/api/notes/v1/list/"
        response = requests.get(api_url)
        response.raise_for_status() # Raise an HTTPError for bad responses (4xx or 5xx)
        api_data = response.json()

        # Insert data into the NOTES sheet
        for record in api_data:
            notes_sheet.append([
                record.get("product_code", ""), # Use .get() for safer access
                record.get("lot_number", ""),
                record.get("product_kind_id", "")
            ])

    except requests.exceptions.RequestException as e:
        Messagebox.show_error(f"Error fetching data from API for NOTES sheet: {e}", "API Error")
        return # Exit function if API call fails

    # Apply formatting
    for col in ["A", "B", "C"]:
        for cell in notes_sheet[col]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    notes_sheet["A4"].font = Font(bold=True)

    # Function to create a WHSE sheet
    def create_whse_sheet(sheet_name):
        sheet = wb.create_sheet(sheet_name)

        # Set header based on warehouse number
        wh_header = "" # Initialize wh_header
        if sheet_name == "WHSE1":
            wh_header = "WHSE #1 - Excess"
        elif sheet_name == "WHSE2":
            wh_header = "WHSE #2 - Excess"
        elif sheet_name == "WHSE4":
            wh_header = "WHSE #4 - Excess"



        # Populate the header (original code had "Date" as first column, but A1 is already date)
        # Assuming "Date" in header means the RM Code or similar, adjusting
        header = [
            "Date", "No of bags", "qty per packing",
            f"{wh_header}", "Total", "Status"
        ]

        sheet.append(header) # This will append to row 2

        # Insert a row for the report date and set its style (original code had A1 as date)
        # To maintain original structure, we'll append header after date if A1 is date
        sheet["A1"] = f"{wh_date}"  # Example date
        sheet["A1"].font = Font(bold=True)
        sheet["B1"].font = Font(bold=True)
        sheet["C1"].font = Font(bold=True)
        sheet["D1"].font = Font(bold=True)
        sheet["E1"].font = Font(bold=True)
        sheet["F1"].font = Font(bold=True)

        # Insert data into the respective warehouse sheet
        for record in data:
            # Ensure 'warehousenumber' exists and matches sheet number
            if record.get("warehousenumber") == int(sheet_name[-1]):
                row = [
                    record.get("rmcode", ""),  # rmcode
                    "",  # No of bags (blank)
                    "",  # qty per packing (blank)
                    "",  # Excess column (blank)
                    float(record.get("new_beginning_balance", 0.0)),  # Total, ensure float
                    "" if record.get("status", "").lower() == "good" else record.get("status", ""),  # Store blank if status is 'good'
                    ""  # Drop list (blank)
                ]
                sheet.append(row)

                # Apply number format (thousands separator)
                # Column 5 is 'Total' (E column)
                qty_cell = sheet.cell(row=sheet.max_row, column=5)
                qty_cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

        # Create a dropdown list for the "drop list" column (Column G)
        dv = DataValidation(
            type="list",
            formula1='"held : under evaluation,held : reject,held : contaminated"',
            allow_blank=True,
            showDropDown=True
        )
        # Apply the data validation to the G column for rows starting from row 3 (after date and header)
        for row_num in range(3, sheet.max_row + 1): # Start from row 3, go to max_row
            cell = f"G{row_num}"
            dv.add(sheet[cell])
        sheet.add_data_validation(dv)

    # Create sheets for WHSE1, WHSE2, and WHSE4
    create_whse_sheet("WHSE1")
    create_whse_sheet("WHSE2")
    create_whse_sheet("WHSE4")

    # Use tkinter's asksaveasfilename for file dialog
    file_path = asksaveasfilename(
        title="Save Excel File",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")]
    )

    # Save the workbook
    if file_path:
        try:
            # Check if the file exists and is open before saving
            if os.path.exists(file_path):
                try:
                    # Attempt to open in exclusive mode to check if it's locked
                    with open(file_path, 'rb+') as f:
                        pass
                except IOError:
                    Messagebox.show_error(
                        f"Error: The file '{os.path.basename(file_path)}' is currently open in another program (e.g., Microsoft Excel). Please close it and try again.",
                        "File In Use"
                    )
                    return # Exit function if file is in use

            # Save the workbook initially using openpyxl (unencrypted)
            wb.save(file_path)

            # --- Apply full workbook password protection using win32com.client ---
            if excel_password and WIN32_AVAILABLE:
                excel_app = None
                try:
                    excel_app = win32.Dispatch("Excel.Application")
                    excel_app.Visible = False # Keep Excel hidden
                    excel_app.DisplayAlerts = False # Suppress alerts

                    # Open the workbook
                    workbook = excel_app.Workbooks.Open(file_path)

                    # Set the password for opening the workbook
                    workbook.Password = excel_password
                    workbook.Save() # Save the workbook with the new password
                    workbook.Close()

                except Exception as e:
                    Messagebox.show_error(
                        f"Error applying workbook password with Excel automation. "
                        f"Ensure Microsoft Excel is installed and pywin32 is correctly installed.\n\nError details: {e}",
                        "Excel Automation Error"
                    )
                finally:
                    if excel_app:
                        excel_app.Quit() # Always quit Excel application
                        del excel_app # Release the COM object
                        time.sleep(0.5) # Give Excel a moment to fully close

            elif excel_password and not WIN32_AVAILABLE:
                Messagebox.show_warning(
                    "Full workbook encryption was requested but 'pywin32' library is not installed or not available. "
                    "The file has been saved, but it is NOT password-protected for opening. "
                    "To enable full workbook encryption, install 'pywin32' using: pip install pywin32",
                    "Pywin32 Not Found"
                )

            Messagebox.show_info(f"Excel file saved successfully at:\n{file_path}", "File Saved")

        except PermissionError as e:
            Messagebox.show_error(
                f"Permission denied: You do not have sufficient permissions to save the file to '{file_path}'.\n\nError details: {e}",
                "Permission Error"
            )
        except IOError as e:
            Messagebox.show_error(
                f"I/O Error: Could not write the file to '{file_path}'. This might be due to an invalid path or disk issues.\n\nError details: {e}",
                "I/O Error"
            )
        except Exception as e:
            Messagebox.show_error(
                f"An unexpected error occurred while saving the file: {e}",
                "Error Saving File"
            )
    else:
        # Saving the file was cancelled
        Messagebox.show_warning("File save cancelled.", "Cancelled")

def get_soh_data():

    """Fetch data from API and format for table rowdata."""
    url = f"{server_ip}/api/get/new_soh/"
    try:
        response = requests.get(url)
        response.raise_for_status()

        data = response.json()
        return data
    except requests.exceptions.RequestException as e:
        return []

    # Return both buttons as a tuple
    return []







